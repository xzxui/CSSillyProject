import configs
import pydantic
import openpyxl
import json
import time
import ModuleLLMQuery

class Comment(pydantic.BaseModel):
    detailed_comment_on_student_performance: str
    custom_error: str = pydantic.Field(..., description="explanation for any fatal error you want to raise. unless a fatal error is what you want to raise, leave this field empty")

def history_to_json(p2e):
    # Read history excel file
    path_to_history = p2e
    wb_obj = openpyxl.load_workbook(path_to_history)
    sheet_obj = wb_obj.active
    # Find all the names of the columns
    keys = []
    right_most_col = 0
    col = 1
    while True:
        val = sheet_obj.cell(row=1, column=col).value
        if val:
            keys.append(val)
        else:
            break
        right_most_col += 1
        col += 1
    # Find all the values under each column, generating a 'dicts'
    # 'dicts' could look like the following example
    # dicts = [{"col1":11, "col2":21}, {"col1":12, "col2":22}]
    dicts = []
    row = 2
    while True:
        dicts.append({})
        all_empty = True
        for col in range(1, right_most_col+1):
            value = sheet_obj.cell(row=row, column=col).value
            if value:
                all_empty = False
            dicts[-1][keys[col-1]] = sheet_obj.cell(row=row, column=col).value
        if all_empty:
            break
        row += 1
    # Return with json's dumps, making it AI-readable
    return json.dumps(dicts, indent=4)

def ProduceFeedbackForStudent(p2e=configs.path_to_excel_of_testing_history):
    """
    Args:
        1. p2e: of <class 'str'>
    Return:
        of <class 'str'>, a summary of a student's performance
    Process:
        use the information stored in an excel file containing the history of tests that the student has taken to ask an AI to give comment
    """
    print("Fetching history")
    history_in_json = history_to_json(p2e)
    print("Done fetching, now asking AI to give comments")
    before = time.time()
    comment = ModuleLLMQuery.LLMQuery(
        [
            {"role": "system", "content": "You are a responsible and experienced teacher who is giving comments on a student's recent performance on exam papers done for practice, and are here to provide a detailed summary of the student's strengths and areas for improvements."},
            {"role": "user", "content": "Provided is the recent performance of the student on practice exam papers, in the format of json:"+history_in_json}
        ],
        response_format=Comment,
        model="gpt-5-mini"
    )
    print(f"AI responded, took {time.time()-before}s")
    # Allowing the AI to determine edge cases
    if comment.custom_error:
        raise RuntimeError(f"The AI raised an error: {comment.custom_error}")
    print("Comment Produced!")
    return comment.detailed_comment_on_student_performance

if __name__ == "__main__":
    with open("test_folder/ModuleProduceFeedbackForStudent/comment1.txt", 'w') as f:
        f.write(ProduceFeedbackForStudent(p2e="test_folder/ModuleProduceFeedbackForStudent/excel_of_testing_history1.xlsx"))
    with open("test_folder/ModuleProduceFeedbackForStudent/comment2.txt", 'w') as f:
        f.write(ProduceFeedbackForStudent(p2e="test_folder/ModuleProduceFeedbackForStudent/excel_of_testing_history2.xlsx"))
