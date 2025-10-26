import configs
import openpyxl
from pathlib import Path

def CreateExcelOfTestingHistory(marking_result_folder_override=None, output_override=None):
    """
    Args:
        1. marking_result_folder_override: optional argument of type string, pretty much explains itself
        2. output_override: optional argument of type string, pretty much explains itself
    Return:
        No return
    Process:
        Create an excel file about the student's testing history, showing the score and grade achieved in each paper, and the AI's comment on the performance of the student, and save to configs.path_to_excel_of_testing_history. Notice that the path of the folder containing all the student's practice paper history is stored in configs.marking_result_folder. Every single file ending with 'xlsx' under that folder should be a marking result save.
    """
    # Create a workbook
    wb = openpyxl.Workbook()
    ws = wb.active

    # Find all the marking results
    marking_result_folder = marking_result_folder_override or configs.marking_result_folder
    xlsx_files = Path(marking_result_folder).glob('*.xlsx')

    # Add all the information needed into the excel of testing history
    rows = [[]]
    # first = [] it is not declared outside because this way when the variable is not created by the iteration, a fatal error occurs in our favor
    for file_path in xlsx_files:
        print(f"Reading: {file_path}")
        new, first = DetermineRowsToAdd(file_path)
        rows.append(new)
    rows[0] = first
    print(rows)
    for row in rows:
        ws.append(row)
    print(f"Added a total of {len(rows)} rows")

    # Save the excel
    wb.save(output_override or configs.path_to_excel_of_testing_history)
    print(f"Excel of testing history saved to {output_override or configs.path_to_excel_of_testing_history}")

def DetermineRowsToAdd(path_to_excel, start_col=9, end_col=15):
    # Load the marking result
    wb_obj = openpyxl.load_workbook(str(path_to_excel))
    sheet_obj = wb_obj.active
    new_row = [path_to_excel.name]
    override_row = ["Exam Paper Name"] # This is the first row of the final excel file saved by this module
    for col in range(start_col, end_col+1):
        new_row.append(sheet_obj.cell(row=2, column=col).value)
        override_row.append(sheet_obj.cell(row=1, column=col).value)
    print(new_row, override_row)
    return new_row, override_row

# For testing
if __name__ == "__main__":
    CreateExcelOfTestingHistory(
        marking_result_folder_override="test_folder/ModuleCreateExcelOfTestingHistory/history1/",
        output_override="test_folder/ModuleCreateExcelOfTestingHistory/excel_of_testing_history1.xlsx"
    )
    CreateExcelOfTestingHistory(
        marking_result_folder_override="test_folder/ModuleCreateExcelOfTestingHistory/history2/",
        output_override="test_folder/ModuleCreateExcelOfTestingHistory/excel_of_testing_history2.xlsx"
    )
    print("Check test_folder/ModuleCreateExcelOfTestingHistory/excel_of_testing_history.xlsx for the output")
