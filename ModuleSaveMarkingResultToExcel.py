import configs
import ModuleCreateExcelOfTestingHistory
import openpyxl

def SaveMarkingResultToExcel(save_path:str, syllabus_code:str, component_number:str, marking_report:list, strengths:str, weaknesses:str, score:int, total_score_avail, grade:str):
    """
    Args:
        1. save_path (<class 'str'>): the path of the excel file to save to
        2. syllabus_code (<class 'str'>)
        3. component_number (<class 'str'>)
        4. marking_report (<class 'list'>): the marking report
        5. strengths (<class 'str'>): the AI's comment on areas the student did well
        6. weaknesses (<class 'str'>): the AI's comment on areas the student did bad
        7. score (<class 'int'>)
        8. total_score_avail (<class 'int'>)
        9. grade (<class 'str'>): the grade received, e.g. A, B, C, D, E, U
    Return: No return
    Process:
        Save the marking result into an excel file, and call ModuleCreateExcelOfTestingHistory.CreateExcelOfTestingHistory
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    
    ws["A1"] = "Question Number"
    ws["B1"] = "Maximum Mark Possible"
    ws["C1"] = "Marks Received"

    ws["I1"] = "Syllabus Code"
    ws["I2"] = syllabus_code
    ws["J1"] = "Component Number"
    ws["J2"] = component_number
    ws["K1"] = "Strengths of the Student"
    ws["K2"] = strengths
    ws["L1"] = "Weaknesses of the Student"
    ws["L2"] = weaknesses
    ws["M1"] = "Total Marks Attained"
    ws["M2"] = score
    ws["N1"] = "Maximum Total Marks Available"
    ws["N2"] = total_score_avail
    ws["O1"] = "Grade Achieved"
    ws["O2"] = grade

    row = 2
    for question_num, max_mark, mark in marking_report:
        ws["A"+str(row)] = question_num
        ws["B"+str(row)] = max_mark
        ws["C"+str(row)] = mark
        row += 1
    wb.save(save_path)
    
    ModuleCreateExcelOfTestingHistory.CreateExcelOfTestingHistory() #updating testing history excel file

    print("Marking result saved to " + save_path, ".")

if __name__ == "__main__":
    testcode: SaveMarkingResultToExcel('test_folder/marking_result_saved.xlsx','0917', '13' ,  [['1', 5, 4], ['2(a)', 10, 4], ['2(b)(i)', 5,4], ['2(b)(ii)',5,5]] , 'Good at algebra', 'Not good at geometry', 114, 200, 'C')
